# Lecture de données de plusieurs fichiers

import openpyxl
import openpyxl.workbook

wb1 = openpyxl.load_workbook("octobre.xlsx", data_only=True) # Permet de récupérer seulement les valeurs et pas les formules
wb2 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)

sheet1 = wb1["Feuil1"]
sheet2 = wb2["Feuil1"]
sheet3 = wb3["Feuil1"]

# Lecture de toutes les ventes pour chaque mois {"Pommes" : (760, 660, 900), "Poires" : (x, x, x)} etc.

# Fonction qui lit les ventes d'un fichier Excel et les ajoute dans un dictionnaire.
def ajouter_data(sheet, d):
    for row in range(2, sheet.max_row):  # On commence à la ligne 2 pour ignorer l'en-tête
        nom_article = sheet.cell(row, 1).value  # Récupération du nom du produit (colonne A)
        if not nom_article:  # Si la cellule est vide, on arrête la lecture
            break
        
        total_ventes = sheet.cell(row, 4).value  # Récupération du total des ventes (colonne D)
        
        # Vérification si l'article existe déjà dans le dictionnaire
        if d.get(nom_article):
            d[nom_article].append(total_ventes)  # Ajoute la valeur aux ventes existantes
        else:
            d[nom_article] = [total_ventes]  # Crée une nouvelle entrée pour l'article

# Initialisation du dictionnaire qui contiendra toutes les ventes
donnees = {}

ajouter_data(sheet1, donnees)
ajouter_data(sheet2, donnees)
ajouter_data(sheet3, donnees)


print(donnees)

# Création d'un fichier de sortie qui regroupe le total des ventes de chaques produits pour chaque mois

wb_sortie = openpyxl.Workbook() # Crée un nouveau fichier Excel vierge
sheet = wb_sortie.active # Récupère la feuille active (par défaut)

sheet["A1"] = "Article"
sheet["B1"] = "Octobre"
sheet["C1"] = "Novembre"
sheet["D1"] = "Decembre"

row = 2 # Commence à la deuxième ligne pour insérer les données sous les en-têtes

for i in donnees.items():
    nom_article = i[0]
    ventes = i[1]
    sheet.cell(row, 1).value = nom_article # Insère le nom du produit (colonne A)

    # Insérer les ventes mensuelles dans les colonnes B, C et D
    for j in range (0, len(ventes)):
        sheet.cell(row, 2+j).value = ventes[j]
    
    row +=1 # Passage à la ligne suivante


# Ajout d'un graphique sur l'évolution des prix

# Création d'une référence pour la plage de données utilisée dans le graphique
chart_ref = openpyxl.chart.Reference(
    sheet,                       # La feuille Excel où se trouvent les données
    min_col=2,                   # Colonne de départ (B) → Première colonne de données numériques
    min_row=2,                   # Ligne de départ (2) → On commence après l'en-tête
    max_col=sheet.max_column,    # Dernière colonne contenant des données
    max_row=2                    # On ne prend que la deuxième ligne (celle des ventes des pommes)
)

# Création d'une série de données pour le graphique
chart_serie = openpyxl.chart.Series(
    chart_ref,                    # Plage de données référencée précédemment
    title="Total ventes €"        # Nom de la série affichée dans la légende du graphique
)

# Création d'un graphique en barres (BarChart)
chart = openpyxl.chart.BarChart()  # Instancie un graphique de type "barres"

chart.title = "Évolution du prix des pommes"  # Titre affiché au-dessus du graphique

# Ajout de la série de données au graphique
chart.append(chart_serie)  # Associe la série de ventes au graphique

sheet.add_chart(chart, "F2")  # Place le graphique à partir de la cellule F2

wb_sortie.save("total_ventes.xlsx")

# Message de confirmation
print("Fichier 'total_ventes.xlsx' créé avec succès !")




