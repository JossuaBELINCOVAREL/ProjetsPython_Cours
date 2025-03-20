# Manipulation de fichiers Excel avec la bibliothèque openpyxl
# fichier .xlsx

import openpyxl

wb = openpyxl.load_workbook("octobre.xlsx") # ouverture du fichier
print(wb.sheetnames)

sheet = wb["Feuil1"] # ouverture de la feuille qui nous intéresse
cellule = sheet.cell(row=4, column=3).value # valeur de la cellule
print(f"La valeur de la cellule est : {cellule}") # affichage de la valeur de la cellule

# Pour récupérer les valeurs de toute une colonne

for row in range(2, 7):
    cellule2 = sheet.cell(row, 2).value
    print(cellule2)